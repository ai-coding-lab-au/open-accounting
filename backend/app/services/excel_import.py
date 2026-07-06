"""Excel / CSV invoice import.

Reads the first sheet of an .xlsx (or a CSV), treats row 1 as headers,
and proposes a mapping from spreadsheet columns → invoice fields. The user
can override the mapping in the UI before confirming the import.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from openpyxl import load_workbook


INVOICE_FIELDS = [
    "direction",
    "contact_name",
    "contact_abn",
    "invoice_number",
    "issue_date",
    "due_date",
    "subtotal",
    "gst_amount",
    "total",
    "currency",
    "notes",
]


# Heuristic header → field mapping. Lower-cased exact-ish match.
_HEADER_HINTS: dict[str, list[str]] = {
    "contact_name": [
        "supplier", "vendor", "bill from", "from", "customer", "client", "bill to",
        "供应商", "客户", "对方", "公司", "name",
    ],
    "contact_abn": ["abn", "tax id", "税号"],
    "invoice_number": [
        "invoice number", "invoice no", "invoice #", "inv no", "inv #", "ref", "reference",
        "票号", "发票号", "发票号码",
    ],
    "issue_date": [
        "date", "invoice date", "issue date", "issued", "tax date",
        "日期", "开票日期",
    ],
    "due_date": ["due date", "due", "payment due", "到期", "到期日"],
    "subtotal": ["subtotal", "net", "amount ex gst", "excl gst", "ex gst", "不含税", "净额"],
    "gst_amount": ["gst", "tax", "gst amount", "税额"],
    "total": ["total", "amount", "incl gst", "gross", "含税", "合计", "总额"],
    "currency": ["currency", "ccy", "币种"],
    "notes": ["notes", "memo", "description", "desc", "备注", "说明"],
    "direction": ["direction", "type", "ap/ar", "方向", "类型"],
}


def _normalise(header: str) -> str:
    return (header or "").strip().lower()


def propose_mapping(headers: list[str]) -> dict[str, int | None]:
    """For each invoice field, return the column index (0-based) we think matches, or None.

    Two-pass to avoid header bleed (e.g. 'total' hint matching 'subtotal'):
    1. Pass 1: exact match wins. A column matched here is locked.
    2. Pass 2: substring fallback for fields still unmatched, skipping locked columns.
    """
    norm = [_normalise(h) for h in headers]
    mapping: dict[str, int | None] = {f: None for f in INVOICE_FIELDS}
    used_cols: set[int] = set()

    # Pass 1 — exact match
    for field, hints in _HEADER_HINTS.items():
        for i, h in enumerate(norm):
            if i in used_cols or not h:
                continue
            if any(hint == h for hint in hints):
                mapping[field] = i
                used_cols.add(i)
                break

    # Pass 2 — substring fallback
    for field, hints in _HEADER_HINTS.items():
        if mapping[field] is not None:
            continue
        for i, h in enumerate(norm):
            if i in used_cols or not h:
                continue
            if any(hint in h for hint in hints):
                mapping[field] = i
                used_cols.add(i)
                break
    return mapping


def _cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip()


def _parse_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except ValueError:
        pass
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d",
        "%d %b %Y", "%d %B %Y",
        "%d-%b-%Y", "%d-%B-%Y",   # 15-Jul-2026
        "%b %d %Y", "%B %d %Y",   # Jul 15 2026
    ):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None  # let the UI flag this row


def _parse_decimal(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return str(Decimal(str(v)).quantize(Decimal("0.01")))
    s = str(v).strip().replace(",", "").replace("$", "").replace("AUD", "").strip()
    if not s:
        return None
    # Accounting-style negatives: "(250.00)" → "-250.00" (common in AU bank
    # exports and ledgers).
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1].strip()
    # Reject scientific notation (e.g. "1e5"): it's never a legitimate bank
    # amount and silently expands to a huge figure. Only digits, a leading sign
    # (CommBank CSV exports amounts as "+5000.00" / "-180.00"), and a single
    # decimal point are allowed.
    if not re.fullmatch(r"[+-]?\d*\.?\d+", s):
        return None
    s = s.lstrip("+")
    try:
        return str(Decimal(s).quantize(Decimal("0.01")))
    except InvalidOperation:
        return None


def _read_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    if not rows:
        return [], []
    headers = [_cell_to_str(c) for c in rows[0]]
    return headers, rows[1:]


def _read_csv(content: bytes) -> tuple[list[str], list[list[Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def parse_spreadsheet(*, content: bytes, filename: str) -> dict[str, Any]:
    """Read file → return headers + sample rows + proposed mapping (no DB writes)."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in {"xlsx", "xlsm"}:
        headers, data_rows = _read_xlsx(content)
    elif ext == "csv":
        headers, data_rows = _read_csv(content)
    else:
        raise ValueError(f"Unsupported spreadsheet format: .{ext} (use .xlsx or .csv)")

    mapping = propose_mapping(headers)

    rows_out: list[dict[str, Any]] = []
    for i, raw in enumerate(data_rows, start=2):  # row 1 was headers
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue
        cells = [_cell_to_str(c) for c in raw]
        rows_out.append({"row_no": i, "cells": cells, "raw": list(raw)})

    return {
        "headers": headers,
        "mapping": mapping,
        "rows": rows_out,
        "field_options": INVOICE_FIELDS,
    }


def apply_mapping(*, raw_row: list[Any], mapping: dict[str, int | None]) -> dict[str, Any]:
    """Materialise one row into a normalised invoice-shaped dict (still pre-save)."""
    out: dict[str, Any] = {}
    for field, idx in mapping.items():
        if idx is None or idx >= len(raw_row):
            out[field] = None
            continue
        v = raw_row[idx]
        if field in {"issue_date", "due_date"}:
            out[field] = _parse_date(v)
        elif field in {"subtotal", "gst_amount", "total"}:
            out[field] = _parse_decimal(v)
        elif field == "direction":
            s = _cell_to_str(v).upper()
            out[field] = s if s in {"AP", "AR"} else None
        else:
            out[field] = _cell_to_str(v) or None
    return out


def apply_mapping_all(*, rows: Iterable[dict[str, Any]], mapping: dict[str, int | None]) -> list[dict[str, Any]]:
    return [
        {"row_no": r["row_no"], "parsed": apply_mapping(raw_row=r["raw"], mapping=mapping)}
        for r in rows
    ]
